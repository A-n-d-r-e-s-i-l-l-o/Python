import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl import load_workbook
import os

class RegistrosExcel(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Registros de notas")
        self.geometry("400x350")
        self.grid_columnconfigure(0,weight=1)
        self.grid_columnconfigure(1,weight=1)
        self.Titulo = tk.Label(self,text="REGISTRO DE NOTAS")
        self.Titulo.grid(column=0,row=0,columnspan=2,padx=5,pady=5,sticky="we")
        textos = [(1,"Nombre del alumno"),(2,"Nota 1"),(3,"Nota 2"),(4,"Nota 3"),(5,"Nota 4"),(6,"Nota 5")]
        for f,t in textos:
            Tex = tk.Label(self, text=t)
            Tex.grid(column=0,row=f,padx=5,pady=5)
        
        self.listaEntradas = []
        entradas = [1,2,3,4,5,6]
        for f in entradas:
            Entrada = tk.Entry(self)
            Entrada.grid(column=1,row=f)
            self.listaEntradas.append(Entrada)

        self.SubirNota = tk.Button(self, text="Subir notas",command=self.GuardarNotas)
        self.SubirNota.grid(column=1,row=7,pady=5) 

    def GuardarNotas(self):
        nombreArchivo = "Registro de notas.xlsx"
        if os.path.exists(nombreArchivo):
            RegistroNotas = load_workbook(nombreArchivo)
            hoja = RegistroNotas.active
        else:
            RegistroNotas = Workbook()
            hoja = RegistroNotas.active
            hoja.title = "Registro de notas"
            hoja["A1"]="Nombre"
            hoja["B1"]="Nota 1"
            hoja["C1"]="Nota 2"
            hoja["D1"]="Nota 3"
            hoja["E1"]="Nota 4"
            hoja["F1"]="Nota 5"
        
        DatosUsuario = []
        for i,entrada in enumerate(self.listaEntradas):
            EntradaUsuario = entrada.get().strip()
            if EntradaUsuario:
                if i == 0:
                    DatosUsuario.append(EntradaUsuario)
                else:
                    try:
                        if "." in EntradaUsuario:
                            nota = float(EntradaUsuario)
                        else:
                            nota = int(EntradaUsuario)
                        DatosUsuario.append(nota)
                    except ValueError:
                        DatosUsuario.append(0)
            else:
                if i == 0:
                    DatosUsuario.append("Sin nombre")
                else:
                    DatosUsuario.append(0)
        hoja.append(DatosUsuario)
        try:
            RegistroNotas.save(nombreArchivo)
            messagebox.showinfo("Exito",f"Las notas de {DatosUsuario[0]} se guardaron exitosamente")
            for limpiar in self.listaEntradas:
                limpiar.delete(0,tk.END)
        except PermissionError:
            messagebox.showerror("Error","No se pudo guardar. Cierra el archivo 'Registro de notas.xlsx' si lo tienes abierto en Excel e intenta de nuevo.")

        
if __name__ == "__main__":
    Registro = RegistrosExcel()
    Registro.mainloop()